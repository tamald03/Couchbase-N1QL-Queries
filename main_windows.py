"""
Couchbase N1QL Query Generator

This script processes Excel files containing entity names and keys,
and generates SELECT and DELETE N1QL queries for Couchbase.
"""

import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass
from typing import List, Dict, Any, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from enum import Enum


# ===============================
# CONSTANTS
# ===============================

REQUIRED_COLUMNS = {"ENTITY_NAME", "KEY"}
EXCEL_EXTENSIONS = (".xlsx", ".xls")
INPUT_SHEET_NAME = "DETAILED_RUN_STATUS"
RTB_PREFIX = "rtb_"
OC_PREFIX = "oc_"
QUERY_SHEET_WIDTHS = {
    "A": 30,  # ENTITY_NAME
    "B": 16,  # NUMBER_OF_KEYS
    "C": 22,  # TOTAL_KEYS_FOR_ENTITY
    "D": 110,  # SELECT_QUERY
    "E": 110,  # DELETE_QUERY
    "F": 40,  # VALIDATION
    "G": 50,  # OPTIMIZATION
    "H": 35,  # PERFORMANCE
}
REPORT_SHEET_WIDTHS = {
    "A": 35,  # FILE_NAME
    "B": 12,  # STATUS
    "C": 14,  # TOTAL_RECORDS
    "D": 10,  # OC_ROWS
    "E": 10,  # RTB_ROWS
    "F": 12,  # SKIPPED_ROWS
    "G": 10,  # OC_GROUPS
    "H": 10,  # RTB_GROUPS
    "I": 45,  # TEMP_FILE
    "J": 20,  # OC_SHEET
    "K": 20,  # RTB_SHEET
    "L": 14,  # READ_TIME_S
    "M": 14,  # SPLIT_TIME_S
    "N": 18,  # QUERY_GEN_TIME_S
    "O": 14,  # FILE_TIME_S
    "P": 80,  # ERROR
}


# ANSI color helpers (enabled only for TTY; disabled by default on Windows)
USE_COLOR = sys.stdout.isatty() and os.name != "nt"
ANSI_RESET = "\033[0m" if USE_COLOR else ""
ANSI_BOLD = "\033[1m" if USE_COLOR else ""
ANSI_RED = "\033[31m" if USE_COLOR else ""
ANSI_GREEN = "\033[32m" if USE_COLOR else ""
ANSI_YELLOW = "\033[33m" if USE_COLOR else ""
ANSI_BLUE = "\033[34m" if USE_COLOR else ""
ANSI_MAGENTA = "\033[35m" if USE_COLOR else ""
ANSI_CYAN = "\033[36m" if USE_COLOR else ""


# ===============================
# CONFIGURATION
# ===============================

# Update these paths as needed (Windows-friendly defaults)
# Example root: r"C:\Data_copy_query_validation_CB"
DEFAULT_INPUT_FOLDER = Path(r"C:\\Users\\tamald\\OneDrive - AMDOCS\\Desktop\\Work\\Data_copy_query_validation_CB\\files")
DEFAULT_OUTPUT_FOLDER = Path(r"C:\\Users\\tamald\\OneDrive - AMDOCS\\Desktop\\Work\\Data_copy_query_validation_CB\\output_excel")
DEFAULT_TEMP_EXCEL_PATH = Path(r"C:\\Users\\tamald\\OneDrive - AMDOCS\\Desktop\\Work\\Data_copy_query_validation_CB\\temp_excel")
EXECUTION_REPORT_PATH = Path(r"C:\\Users\\tamald\\OneDrive - AMDOCS\\Desktop\\Work\\Data_copy_query_validation_CB\\execution_report")
LOG_FILE_LOCATION = Path(r"C:\\Users\\tamald\\OneDrive - AMDOCS\\Desktop\\Work\\Data_copy_query_validation_CB\\log_files")


@dataclass
class Config:
    input_folder: Path = DEFAULT_INPUT_FOLDER
    output_folder: Path = DEFAULT_OUTPUT_FOLDER
    temp_excel_path: Path = DEFAULT_TEMP_EXCEL_PATH
    execution_report_path: Path = EXECUTION_REPORT_PATH
    log_file_location: Path = LOG_FILE_LOCATION
    input_sheet_name: str = INPUT_SHEET_NAME
    max_keys_per_group: int = 500
    save_temp_sheets: bool = True
    validate_queries: bool = True
    analyze_optimization: bool = True


# ===============================
# HELPER FUNCTIONS
# ===============================

def color(text: str, color_code: str) -> str:
    """Wrap text with ANSI color codes when enabled."""
    if not USE_COLOR:
        return text
    return f"{color_code}{text}{ANSI_RESET}"


def banner(title: str) -> None:
    """Print a section banner to make logs easier to scan."""
    print(color("\n" + "=" * 70, ANSI_CYAN))
    print(color(title, ANSI_BOLD + ANSI_CYAN))
    print(color("=" * 70, ANSI_CYAN))


def step_title(title: str) -> None:
    """Print a step header for progress logging."""
    print(color(f"\n{title}", ANSI_BOLD + ANSI_BLUE))


class StripAnsiStream:
    """Write stream wrapper that strips ANSI sequences."""
    def __init__(self, stream):
        self.stream = stream
        self._ansi_re = re.compile(r"\x1b\[[0-9;]*m")

    def write(self, data):
        cleaned = self._ansi_re.sub("", data)
        self.stream.write(cleaned)

    def flush(self):
        self.stream.flush()

    def isatty(self):
        return False


class Tee:
    """Write stream that duplicates output to multiple streams."""
    def __init__(self, *streams):
        self.streams = streams

    def write(self, data):
        for stream in self.streams:
            stream.write(data)
            stream.flush()

    def flush(self):
        for stream in self.streams:
            stream.flush()

    def isatty(self):
        return any(getattr(stream, "isatty", lambda: False)() for stream in self.streams)


def truncate_cell(text: str, max_width: int) -> str:
    """Shorten long text to fit console table columns."""
    if len(text) <= max_width:
        return text
    if max_width <= 3:
        return text[:max_width]
    return text[: max_width - 3] + "..."


def print_summary_table(report_rows: List[Dict[str, Any]]) -> None:
    """Print a compact execution summary table to the console."""
    headers = ["FILE", "STATUS", "RECORDS", "OC", "RTB", "SKIP", "SECONDS"]
    rows = []
    for row in report_rows:
        rows.append([
            str(row.get("file_name", "")),
            str(row.get("status", "")),
            str(row.get("total_records", 0)),
            str(row.get("oc_rows", 0)),
            str(row.get("rtb_rows", 0)),
            str(row.get("skipped_rows", 0)),
            f"{row.get('file_time_s', 0.0):.2f}",
        ])

    max_widths = [40, 10, 8, 8, 8, 8, 10]
    col_widths = []
    for i, header in enumerate(headers):
        width = max(len(header), max((len(r[i]) for r in rows), default=0))
        col_widths.append(min(width, max_widths[i]))

    def format_row(values: List[str]) -> str:
        cells = []
        for i, val in enumerate(values):
            cell = truncate_cell(val, col_widths[i]).ljust(col_widths[i])
            cells.append(cell)
        return "| " + " | ".join(cells) + " |"

    line = "+" + "+".join(["-" * (w + 2) for w in col_widths]) + "+"

    print(color("\nExecution Summary", ANSI_BOLD + ANSI_CYAN))
    print(color(line, ANSI_CYAN))
    print(color(format_row(headers), ANSI_CYAN))
    print(color(line, ANSI_CYAN))
    for r in rows:
        status = r[1].lower()
        row_text = format_row(r)
        if status == "error":
            print(color(row_text, ANSI_RED))
        elif status == "no data":
            print(color(row_text, ANSI_YELLOW))
        else:
            print(color(row_text, ANSI_GREEN))
    print(color(line, ANSI_CYAN))

def ensure_folders(config: Config) -> bool:
    """
    Validate that all required folders exist (optionally create them).
    
    Returns:
        bool: True if all folders exist, False otherwise
    """
    folders = {
        "INPUT_FOLDER": config.input_folder,
        "OUTPUT_FOLDER": config.output_folder,
        "TEMP_EXCEL_PATH": config.temp_excel_path,
        "EXECUTION_REPORT_PATH": config.execution_report_path,
        "LOG_FILE_LOCATION": config.log_file_location,
    }
    
    all_exist = True
    for name, folder in folders.items():
        if not folder.exists():
            if name == "INPUT_FOLDER":
                print(color(f"ERROR: {name} does not exist: {folder}", ANSI_RED))
                all_exist = False
                continue

            try:
                folder.mkdir(parents=True, exist_ok=True)
                print(color(f"✓ Created {name}: {folder}", ANSI_GREEN))
            except Exception as e:
                print(color(f"ERROR creating {name}: {folder} ({str(e)})", ANSI_RED))
                all_exist = False
        else:
            print(color(f"✓ {name}: {folder}", ANSI_GREEN))
    
    return all_exist


def get_excel_files(folder: Path) -> List[Path]:
    """
    Get all Excel files from the specified folder.
    
    Args:
        folder: Path to the folder containing Excel files
        
    Returns:
        List of Path objects pointing to Excel files
    """
    excel_files = []
    
    if not folder.exists():
        print(color(f"ERROR: Folder does not exist: {folder}", ANSI_RED))
        return excel_files
    
    for file in folder.iterdir():
        if file.is_file() and file.suffix in EXCEL_EXTENSIONS:
            # Skip temporary files
            if not file.name.startswith("~$"):
                excel_files.append(file)
    
    return sorted(excel_files, key=lambda p: p.name.lower())


def read_detailed_run_status(excel_file: Path, sheet_name: str) -> pd.DataFrame:
    """
    Read the DETAILED_RUN_STATUS sheet from an Excel file.
    
    Args:
        excel_file: Path to the Excel file
        
    Returns:
        List of rows with ENTITY_NAME and KEY
    """
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, engine="openpyxl")

        # Normalize column names to improve matching
        df.columns = [str(col).strip() for col in df.columns]
        col_map = {str(col).strip().lower(): col for col in df.columns}

        required_lower = {c.lower() for c in REQUIRED_COLUMNS}
        missing_lower = required_lower - set(col_map.keys())
        if missing_lower:
            print(color(
                f"WARNING: Missing columns in {excel_file.name}: "
                f"{sorted(missing_lower)} | Available: {list(df.columns)}",
                ANSI_YELLOW,
            ))
            return pd.DataFrame(columns=list(REQUIRED_COLUMNS))

        actual_entity = col_map["entity_name"]
        actual_key = col_map["key"]

        # Keep only required columns and standardize names
        df = df[[actual_entity, actual_key]].copy()
        df.columns = ["ENTITY_NAME", "KEY"]

        # Remove rows with missing values
        df = df.dropna()

        return df
        
    except Exception as e:
        print(color(f"ERROR reading {excel_file.name}: {str(e)}", ANSI_RED))
        return pd.DataFrame(columns=list(REQUIRED_COLUMNS))


def split_entities_by_prefix_rowwise(
    df: pd.DataFrame,
) -> Tuple[Dict[str, pd.DataFrame], Dict[str, int]]:
    """
    Split rows into OC and RTB DataFrames by ENTITY_NAME prefix.

    This preserves the original row order from the input sheet.
    """
    def normalize_entity_name(value: Any) -> str:
        text = str(value) if value is not None else ""
        text = text.replace("\u00a0", " ").strip()
        # Drop leading non-alnum chars (handles hidden/bom-like prefixes)
        while text and not text[0].isalnum():
            text = text[1:]
        return text.strip()

    oc_rows = []
    rtb_rows = []
    skipped = 0

    for row in df.itertuples(index=False):
        entity_name = getattr(row, "ENTITY_NAME", None)
        key_value = getattr(row, "KEY", None)

        entity_clean = normalize_entity_name(entity_name)
        if not entity_clean:
            skipped += 1
            continue

        entity_lower = entity_clean.lower()
        if entity_lower.startswith(OC_PREFIX):
            oc_rows.append({"ENTITY_NAME": format_entity_name(entity_clean), "KEY": key_value})
        elif entity_lower.startswith(RTB_PREFIX):
            rtb_rows.append({"ENTITY_NAME": format_entity_name(entity_clean), "KEY": key_value})
        else:
            skipped += 1

    print(color(
        f"  Rows routed → OC: {len(oc_rows)}, RTB: {len(rtb_rows)}, Skipped: {skipped}",
        ANSI_MAGENTA,
    ))

    return {
        "oc": pd.DataFrame(oc_rows, columns=["ENTITY_NAME", "KEY"]),
        "rtb": pd.DataFrame(rtb_rows, columns=["ENTITY_NAME", "KEY"]),
    }, {
        "oc_rows": len(oc_rows),
        "rtb_rows": len(rtb_rows),
        "skipped_rows": skipped,
    }


def save_temp_sheets(oc_df: pd.DataFrame, rtb_df: pd.DataFrame, temp_file: Path) -> None:
    """
    Save OC and RTB rows to temporary Excel file.
    
    Args:
        oc_df: DataFrame containing OC entities
        rtb_df: DataFrame containing RTB entities
        temp_file: Path to the temporary Excel file
    """
    try:
        workbook = Workbook()
        oc_sheet = workbook.active
        oc_sheet.title = "OC"
        rtb_sheet = workbook.create_sheet(title="RTB")

        oc_sheet.append(["ENTITY_NAME", "KEY"])
        for row in oc_df.itertuples(index=False):
            oc_sheet.append([getattr(row, "ENTITY_NAME", None), getattr(row, "KEY", None)])

        rtb_sheet.append(["ENTITY_NAME", "KEY"])
        for row in rtb_df.itertuples(index=False):
            rtb_sheet.append([getattr(row, "ENTITY_NAME", None), getattr(row, "KEY", None)])

        workbook.save(temp_file)

        print(color(f"  Saved temporary file: {temp_file.name}", ANSI_GREEN))
        print(color(f"    OC entities: {len(oc_df)}", ANSI_CYAN))
        print(color(f"    RTB entities: {len(rtb_df)}", ANSI_CYAN))
        
    except Exception as e:
        print(color(f"ERROR saving temporary file {temp_file.name}: {str(e)}", ANSI_RED))


def group_keys_by_entity(df: pd.DataFrame, max_keys: int = 500) -> List[Dict[str, Any]]:
    """
    Group keys by entity name and split if more than max_keys.
    
    Args:
        df: DataFrame containing ENTITY_NAME and KEY columns
        max_keys: Maximum number of keys per group (default: 500)
        
    Returns:
        List of dictionaries with entity_name, keys, total_keys, part_number
    """
    if df.empty:
        return []
    
    # Sort by entity name and key
    result = []
    
    df_sorted = df.sort_values(["ENTITY_NAME", "KEY"]).reset_index(drop=True)
    grouped = df_sorted.groupby("ENTITY_NAME")

    for entity_name, group in grouped:
        keys = group["KEY"].tolist()
        total_keys = len(keys)
        
        # Split into chunks if needed
        if total_keys > max_keys:
            num_parts = (total_keys + max_keys - 1) // max_keys  # Ceiling division
            
            for part_num in range(num_parts):
                start_idx = part_num * max_keys
                end_idx = min((part_num + 1) * max_keys, total_keys)
                key_chunk = keys[start_idx:end_idx]
                
                result.append({
                    'entity_name': entity_name,
                    'keys': key_chunk,
                    'total_keys': total_keys,
                    'part_number': part_num + 1,
                    'total_parts': num_parts,
                    'chunk_size': len(key_chunk)
                })
        else:
            result.append({
                'entity_name': entity_name,
                'keys': keys,
                'total_keys': total_keys,
                'part_number': 1,
                'total_parts': 1,
                'chunk_size': total_keys
            })
    
    return result


def generate_queries(entity_name: str, keys: List[str], config: Config) -> Dict[str, Any]:
    """
    Generate SELECT and DELETE N1QL queries for an entity with keys.
    Includes validation, optimization analysis, and performance estimation.
    
    Args:
        entity_name: Name of the entity/bucket
        keys: List of keys
        config: Configuration object
        
    Returns:
        Dictionary with queries and analysis data
    """
    # Convert keys to JSON array format
    keys_str = str(keys).replace("'", '"')
    
    entity_ref = format_entity_name(entity_name)
    select_query = (
        "SELECT\n"
        "    k AS key_name,\n"
        "    META(e).id AS document_id,\n"
        "    CASE\n"
        "        WHEN e IS MISSING THEN \"Not Found\"\n"
        "        ELSE \"Found\"\n"
        "    END AS status\n"
        f"FROM {keys_str} AS k\n"
        f"LEFT JOIN {entity_ref} e ON KEYS k;"
    )
    delete_query = f"DELETE FROM {entity_ref} USE KEYS {keys_str};"
    
    result = {
        'select_query': select_query,
        'delete_query': delete_query,
        'validation_issues': [],
        'optimization_suggestions': [],
        'performance_estimate': None
    }
    
    # Validate queries if enabled
    if config.validate_queries:
        select_issues = validate_n1ql_query(select_query, 'SELECT')
        delete_issues = validate_n1ql_query(delete_query, 'DELETE')
        result['validation_issues'] = select_issues + delete_issues
    
    # Analyze optimization if enabled
    if config.analyze_optimization:
        result['optimization_suggestions'] = analyze_query_optimization(
            select_query, len(keys), entity_name
        )
    
    # Estimate performance if enabled
    if config.analyze_optimization:
        result['performance_estimate'] = estimate_query_performance(
            select_query, len(keys), 'SELECT'
        )
    
    return result


def create_output_dataframe(entity_groups: List[Dict[str, Any]], config: Config) -> pd.DataFrame:
    """
    Create output rows with entity information and queries.
    
    Args:
        entity_groups: List of entity group dictionaries
        config: Configuration object
        
    Returns:
        DataFrame with ENTITY_NAME, NUMBER_OF_KEYS, SELECT_QUERY, DELETE_QUERY and analysis data
    """
    output_data = []
    
    for group in entity_groups:
        query_result = generate_queries(group['entity_name'], group['keys'], config)
        
        # Create entity name with part info if split
        entity_display = format_entity_name(group['entity_name'])
        if group['total_parts'] > 1:
            entity_display = f"{entity_display} (Part {group['part_number']}/{group['total_parts']})"
        
        # Build validation summary
        validation_summary = ""
        if query_result['validation_issues']:
            errors = [v for v in query_result['validation_issues'] if v.level == QueryValidationResult.ERROR]
            warnings = [v for v in query_result['validation_issues'] if v.level == QueryValidationResult.WARNING]
            
            if errors:
                validation_summary += f"❌ {len(errors)} Error(s): " + "; ".join([v.message for v in errors]) + "\n"
            if warnings:
                validation_summary += f"⚠ {len(warnings)} Warning(s): " + "; ".join([v.message for v in warnings])
        else:
            validation_summary = "✅ Valid"
        
        # Build optimization summary
        optimization_summary = ""
        if query_result['optimization_suggestions']:
            high_impact = [s for s in query_result['optimization_suggestions'] if s.impact == "High"]
            if high_impact:
                optimization_summary = f"🔴 {len(high_impact)} High Impact | " + high_impact[0].suggestion[:50] + "..."
            else:
                optimization_summary = f"✅ {len(query_result['optimization_suggestions'])} suggestions"
        
        # Build performance summary
        performance_summary = ""
        if query_result['performance_estimate']:
            perf = query_result['performance_estimate']
            performance_summary = f"{perf.complexity} | ~{perf.estimated_time_ms:.0f}ms | {perf.key_count} keys"
        
        output_data.append({
            'ENTITY_NAME': entity_display,
            'NUMBER_OF_KEYS': group['chunk_size'],
            'TOTAL_KEYS_FOR_ENTITY': group['total_keys'],
            'SELECT_QUERY': query_result['select_query'],
            'DELETE_QUERY': query_result['delete_query'],
            'VALIDATION': validation_summary,
            'OPTIMIZATION': optimization_summary,
            'PERFORMANCE': performance_summary
        })
    
    return pd.DataFrame(output_data)


def save_output_file(df: pd.DataFrame, output_file: Path) -> None:
    """
    Save output rows to Excel file.
    
    Args:
        df: DataFrame to save
        output_file: Path to the output Excel file
    """
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Queries"

        headers = ["ENTITY_NAME", "NUMBER_OF_KEYS", "TOTAL_KEYS_FOR_ENTITY", "SELECT_QUERY", "DELETE_QUERY", 
                   "VALIDATION", "OPTIMIZATION", "PERFORMANCE"]
        sheet.append(headers)
        for row in df.itertuples(index=False):
            sheet.append([
                getattr(row, "ENTITY_NAME", None),
                getattr(row, "NUMBER_OF_KEYS", None),
                getattr(row, "TOTAL_KEYS_FOR_ENTITY", None),
                getattr(row, "SELECT_QUERY", None),
                getattr(row, "DELETE_QUERY", None),
                getattr(row, "VALIDATION", ""),
                getattr(row, "OPTIMIZATION", ""),
                getattr(row, "PERFORMANCE", ""),
            ])

        apply_query_sheet_formatting(sheet)
        workbook.save(output_file)
        print(color(f"  ✓ Saved output file: {output_file.name}", ANSI_GREEN))
        print(color(f"    Total query groups: {len(df)}", ANSI_CYAN))
        
    except Exception as e:
        print(color(f"ERROR saving output file {output_file.name}: {str(e)}", ANSI_RED))


def write_df_to_sheet(workbook: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Write a DataFrame to a worksheet with a standard header.
    """
    sheet = workbook.create_sheet(title=sheet_name)
    headers = ["ENTITY_NAME", "NUMBER_OF_KEYS", "TOTAL_KEYS_FOR_ENTITY", "SELECT_QUERY", "DELETE_QUERY",
               "VALIDATION", "OPTIMIZATION", "PERFORMANCE"]
    sheet.append(headers)
    for row in df.itertuples(index=False):
        sheet.append([
            getattr(row, "ENTITY_NAME", None),
            getattr(row, "NUMBER_OF_KEYS", None),
            getattr(row, "TOTAL_KEYS_FOR_ENTITY", None),
            getattr(row, "SELECT_QUERY", None),
            getattr(row, "DELETE_QUERY", None),
            getattr(row, "VALIDATION", ""),
            getattr(row, "OPTIMIZATION", ""),
            getattr(row, "PERFORMANCE", ""),
        ])
    apply_query_sheet_formatting(sheet)


def apply_query_sheet_formatting(sheet) -> None:
    """
    Apply column widths and wrapping for query sheets.
    """
    for col, width in QUERY_SHEET_WIDTHS.items():
        sheet.column_dimensions[col].width = width

    wrap = Alignment(wrap_text=True, vertical="top")
    # Apply wrapping to query columns (D, E) and analysis columns (F, G, H)
    for row in sheet.iter_rows(min_row=1, min_col=4, max_col=8):
        for cell in row:
            cell.alignment = wrap


def make_sheet_title(base_name: str, suffix: str, existing: set) -> str:
    """
    Build a safe, unique sheet name (max 31 chars).
    """
    raw = f"{base_name}_{suffix}".replace(" ", "_")
    base_title = raw[:31]
    if base_title not in existing:
        existing.add(base_title)
        return base_title

    for i in range(1, 1000):
        tail = f"_{i}"
        candidate = f"{base_title[:31 - len(tail)]}{tail}"
        if candidate not in existing:
            existing.add(candidate)
            return candidate

    raise ValueError("Could not generate unique sheet name.")


def save_execution_report(report_rows: List[Dict[str, Any]], report_file: Path) -> None:
    """
    Save a single execution summary report for all input files.
    """
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "ExecutionReport"

        headers = [
            "FILE_NAME",
            "STATUS",
            "TOTAL_RECORDS",
            "OC_ROWS",
            "RTB_ROWS",
            "SKIPPED_ROWS",
            "OC_GROUPS",
            "RTB_GROUPS",
            "TEMP_FILE",
            "OC_SHEET",
            "RTB_SHEET",
            "READ_TIME_S",
            "SPLIT_TIME_S",
            "QUERY_GEN_TIME_S",
            "FILE_TIME_S",
            "ERROR",
        ]
        sheet.append(headers)

        for row in report_rows:
            sheet.append([
                row.get("file_name", ""),
                row.get("status", ""),
                row.get("total_records", 0),
                row.get("oc_rows", 0),
                row.get("rtb_rows", 0),
                row.get("skipped_rows", 0),
                row.get("oc_groups", 0),
                row.get("rtb_groups", 0),
                row.get("temp_file", ""),
                row.get("oc_sheet", ""),
                row.get("rtb_sheet", ""),
                row.get("read_time_s", 0.0),
                row.get("split_time_s", 0.0),
                row.get("query_gen_time_s", 0.0),
                row.get("file_time_s", 0.0),
                row.get("error", ""),
            ])

        apply_report_sheet_formatting(sheet)
        workbook.save(report_file)
        print(color(f"\n✓ Execution report saved to: {report_file}", ANSI_GREEN))
    except Exception as e:
        print(color(f"ERROR saving execution report {report_file}: {str(e)}", ANSI_RED))


def process_excel_file(
    excel_file: Path,
    config: Config,
) -> Tuple[Dict[str, Any], pd.DataFrame, pd.DataFrame]:
    """
    Process a single Excel file through the complete workflow.
    
    Args:
        excel_file: Path to the Excel file to process
    """
    file_start = time.time()
    banner(f"Processing: {excel_file.name}")
    
    # Step 1: Read DETAILED_RUN_STATUS sheet
    step_title("[Step 1] Reading DETAILED_RUN_STATUS sheet...")
    read_start = time.time()
    df = read_detailed_run_status(excel_file, config.input_sheet_name)
    read_time = time.time() - read_start
    
    if df.empty:
        print(color("  No data found. Skipping file.", ANSI_YELLOW))
        print(color(f"  Finished in {time.time() - file_start:.2f}s", ANSI_CYAN))
        file_time = time.time() - file_start
        return {
            "file_name": excel_file.name,
            "status": "No data",
            "total_records": 0,
            "oc_rows": 0,
            "rtb_rows": 0,
            "skipped_rows": 0,
            "oc_groups": 0,
            "rtb_groups": 0,
            "temp_file": "",
            "oc_sheet": "",
            "rtb_sheet": "",
            "read_time_s": round(read_time, 2),
            "split_time_s": 0.0,
            "query_gen_time_s": 0.0,
            "file_time_s": round(file_time, 2),
            "error": "",
        }, pd.DataFrame(), pd.DataFrame()
    
    print(color(f"  Total records: {len(df)}", ANSI_CYAN))
    
    # Step 2: Split by OC and RTB prefix
    step_title("[Step 2] Splitting entities by prefix...")
    split_start = time.time()
    split_data, split_stats = split_entities_by_prefix_rowwise(df)
    split_time = time.time() - split_start
    
    # Save temporary sheets
    temp_file = ""
    if config.save_temp_sheets:
        temp_path = config.temp_excel_path / f"temp_{excel_file.stem}.xlsx"
        save_temp_sheets(split_data['oc'], split_data['rtb'], temp_path)
        temp_file = str(temp_path)
    
    # Step 3 & 4: Group keys by entity and split if needed
    step_title("[Step 3-4] Grouping and splitting entities...")
    
    oc_groups = group_keys_by_entity(split_data['oc'], max_keys=config.max_keys_per_group)
    rtb_groups = group_keys_by_entity(split_data['rtb'], max_keys=config.max_keys_per_group)
    
    print(color(f"  OC entity groups: {len(oc_groups)}", ANSI_CYAN))
    print(color(f"  RTB entity groups: {len(rtb_groups)}", ANSI_CYAN))
    
    # Show entities with splits
    for groups, prefix in [(oc_groups, 'OC'), (rtb_groups, 'RTB')]:
        split_entities = [g for g in groups if g['total_parts'] > 1]
        if split_entities:
            print(color(f"\n  {prefix} entities split into parts:", ANSI_YELLOW))
            for g in split_entities:
                print(color(
                    f"    - {g['entity_name']}: {g['total_keys']} keys → {g['total_parts']} parts",
                    ANSI_YELLOW,
                ))
    
    # Step 5: Generate queries and create output DataFrames
    step_title("[Step 5] Generating queries with validation and analysis...")
    query_start = time.time()
    oc_output_df = create_output_dataframe(oc_groups, config)
    rtb_output_df = create_output_dataframe(rtb_groups, config)
    query_gen_time = time.time() - query_start
    
    # Print analysis summary
    if config.validate_queries or config.analyze_optimization:
        print(color(f"  ✓ Validation: {'Enabled' if config.validate_queries else 'Disabled'}", ANSI_CYAN))
        print(color(f"  ✓ Optimization Analysis: {'Enabled' if config.analyze_optimization else 'Disabled'}", ANSI_CYAN))
    
    # Step 6: Save output files
    step_title("[Step 6] Saving output files...")
    
    if oc_output_df.empty:
        print(color("  No OC entities to save", ANSI_YELLOW))
    if rtb_output_df.empty:
        print(color("  No RTB entities to save", ANSI_YELLOW))
    
    file_time = time.time() - file_start
    banner(f"Completed: {excel_file.name} in {file_time:.2f}s")

    return {
        "file_name": excel_file.name,
        "status": "Processed",
        "total_records": len(df),
        "oc_rows": split_stats["oc_rows"],
        "rtb_rows": split_stats["rtb_rows"],
        "skipped_rows": split_stats["skipped_rows"],
        "oc_groups": len(oc_groups),
        "rtb_groups": len(rtb_groups),
        "temp_file": temp_file,
        "oc_sheet": "",
        "rtb_sheet": "",
        "read_time_s": round(read_time, 2),
        "split_time_s": round(split_time, 2),
        "query_gen_time_s": round(query_gen_time, 2),
        "file_time_s": round(file_time, 2),
        "error": "",
    }, oc_output_df, rtb_output_df


def apply_report_sheet_formatting(sheet) -> None:
    """
    Apply column widths and wrapping for execution report sheets.
    """
    for col, width in REPORT_SHEET_WIDTHS.items():
        sheet.column_dimensions[col].width = width

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows(min_row=1, min_col=len(REPORT_SHEET_WIDTHS), max_col=len(REPORT_SHEET_WIDTHS)):
        for cell in row:
            cell.alignment = wrap


class QueryValidationResult(Enum):
    VALID = "Valid"
    WARNING = "Warning"
    ERROR = "Error"

@dataclass
class ValidationIssue:
    level: QueryValidationResult
    message: str
    query_type: str

def validate_n1ql_query(query: str, query_type: str) -> List[ValidationIssue]:
    """
    Comprehensive N1QL query validation.
    
    Args:
        query: The N1QL query string to validate
        query_type: Type of query ('SELECT' or 'DELETE')
        
    Returns:
        List of ValidationIssue objects
    """
    issues = []
    query_upper = query.upper()
    
    # Basic bracket matching
    if query.count('[') != query.count(']'):
        issues.append(ValidationIssue(
            QueryValidationResult.ERROR, 
            "Unbalanced square brackets []", 
            query_type
        ))
    
    if query.count('(') != query.count(')'):
        issues.append(ValidationIssue(
            QueryValidationResult.ERROR, 
            "Unbalanced parentheses ()", 
            query_type
        ))
    
    # Quote matching
    if query.count('"') % 2 != 0:
        issues.append(ValidationIssue(
            QueryValidationResult.WARNING, 
            "Unbalanced double quotes", 
            query_type
        ))
    
    # Query-type specific validation
    if query_type == 'SELECT':
        if 'FROM' not in query_upper:
            issues.append(ValidationIssue(
                QueryValidationResult.ERROR, 
                "SELECT query missing FROM clause", 
                query_type
            ))
        
        if 'SELECT' not in query_upper:
            issues.append(ValidationIssue(
                QueryValidationResult.ERROR, 
                "Missing SELECT keyword", 
                query_type
            ))
        
        # Check for potential issues with LEFT JOIN
        if 'LEFT JOIN' in query_upper and 'ON KEYS' not in query_upper:
            issues.append(ValidationIssue(
                QueryValidationResult.WARNING, 
                "LEFT JOIN without ON KEYS clause", 
                query_type
            ))
    
    elif query_type == 'DELETE':
        if 'DELETE' not in query_upper:
            issues.append(ValidationIssue(
                QueryValidationResult.ERROR, 
                "Missing DELETE keyword", 
                query_type
            ))
        
        if 'FROM' not in query_upper:
            issues.append(ValidationIssue(
                QueryValidationResult.ERROR, 
                "DELETE query missing FROM clause", 
                query_type
            ))
        
        if 'USE KEYS' not in query_upper and 'WHERE' not in query_upper:
            issues.append(ValidationIssue(
                QueryValidationResult.WARNING, 
                "DELETE without USE KEYS or WHERE clause (potential safety issue)", 
                query_type
            ))
    
    # Check for common syntax errors
    if query_upper.count('SELECT') > 1 and 'UNION' not in query_upper:
        issues.append(ValidationIssue(
            QueryValidationResult.WARNING, 
            "Multiple SELECT keywords without UNION", 
            query_type
        ))
    
    # Check for semicolon at end
    if not query.rstrip().endswith(';'):
        issues.append(ValidationIssue(
            QueryValidationResult.WARNING, 
            "Query does not end with semicolon", 
            query_type
        ))
    
    # Check for potential SQL injection patterns (basic check)
    dangerous_patterns = ['--', '/*', '*/']
    for pattern in dangerous_patterns:
        if pattern in query and pattern != '-- ':
            issues.append(ValidationIssue(
                QueryValidationResult.WARNING, 
                f"Potential SQL comment pattern detected: {pattern}", 
                query_type
            ))
    
    return issues


@dataclass
class OptimizationSuggestion:
    category: str
    suggestion: str
    impact: str  # "High", "Medium", "Low"
    current_issue: str


@dataclass
class PerformanceEstimate:
    estimated_time_ms: float
    complexity: str  # "Simple", "Moderate", "Complex"
    key_count: int
    factors: List[str]


def analyze_query_optimization(query: str, key_count: int, entity_name: str) -> List[OptimizationSuggestion]:
    """
    Analyze N1QL query and provide optimization suggestions.
    
    Args:
        query: The N1QL query string
        key_count: Number of keys in the query
        entity_name: Entity/bucket name
        
    Returns:
        List of OptimizationSuggestion objects
    """
    suggestions = []
    query_upper = query.upper()
    
    # Check for USE KEYS optimization
    if 'USE KEYS' in query_upper:
        suggestions.append(OptimizationSuggestion(
            category="Index Usage",
            suggestion="Query uses USE KEYS which provides direct key lookup - optimal performance",
            impact="High",
            current_issue="None - already optimized"
        ))
    else:
        suggestions.append(OptimizationSuggestion(
            category="Index Usage",
            suggestion="Consider using USE KEYS for direct document access instead of WHERE clause",
            impact="High",
            current_issue="Missing USE KEYS clause for key-based lookup"
        ))
    
    # Check for SELECT *
    if 'SELECT *' in query_upper or 'SELECT\n*' in query_upper:
        suggestions.append(OptimizationSuggestion(
            category="Projection",
            suggestion="Avoid SELECT * - specify only required fields to reduce data transfer",
            impact="Medium",
            current_issue="Using SELECT * retrieves all fields"
        ))
    
    # Check for large batch size
    if key_count > 1000:
        suggestions.append(OptimizationSuggestion(
            category="Batch Size",
            suggestion=f"Key count ({key_count}) is large. Consider splitting into smaller batches of 500-1000 keys",
            impact="High",
            current_issue=f"Large batch size may cause timeout or memory issues"
        ))
    elif key_count > 500:
        suggestions.append(OptimizationSuggestion(
            category="Batch Size",
            suggestion=f"Key count ({key_count}) is moderate. Monitor query performance",
            impact="Medium",
            current_issue="Batch size approaching recommended limit"
        ))
    
    # Check for JOIN operations
    if 'JOIN' in query_upper:
        if 'ON KEYS' not in query_upper:
            suggestions.append(OptimizationSuggestion(
                category="JOIN Performance",
                suggestion="Use ON KEYS for JOIN operations instead of ON clause for better performance",
                impact="High",
                current_issue="JOIN without ON KEYS may require index scan"
            ))
        else:
            suggestions.append(OptimizationSuggestion(
                category="JOIN Performance",
                suggestion="JOIN with ON KEYS is efficient for key-based lookups",
                impact="Low",
                current_issue="None - JOIN is optimized"
            ))
    
    # Check for subqueries
    if query_upper.count('SELECT') > 1:
        suggestions.append(OptimizationSuggestion(
            category="Query Complexity",
            suggestion="Subqueries detected. Consider flattening or using JOINs if possible",
            impact="Medium",
            current_issue="Subqueries may impact performance"
        ))
    
    # Check for ORDER BY without LIMIT
    if 'ORDER BY' in query_upper and 'LIMIT' not in query_upper:
        suggestions.append(OptimizationSuggestion(
            category="Result Set",
            suggestion="ORDER BY without LIMIT sorts entire result set. Add LIMIT if possible",
            impact="Medium",
            current_issue="Sorting without limiting result size"
        ))
    
    # Check for DISTINCT
    if 'DISTINCT' in query_upper:
        suggestions.append(OptimizationSuggestion(
            category="Data Processing",
            suggestion="DISTINCT requires additional processing. Ensure it's necessary",
            impact="Low",
            current_issue="DISTINCT operation adds overhead"
        ))
    
    return suggestions


def estimate_query_performance(query: str, key_count: int, query_type: str) -> PerformanceEstimate:
    """
    Estimate query performance based on query characteristics.
    
    Args:
        query: The N1QL query string
        key_count: Number of keys
        query_type: Type of query ('SELECT' or 'DELETE')
        
    Returns:
        PerformanceEstimate object
    """
    query_upper = query.upper()
    factors = []
    base_time_ms = 0.0
    
    # Base time for key lookup
    if 'USE KEYS' in query_upper:
        base_time_ms = 1.0 * key_count  # ~1ms per key for direct lookup
        factors.append(f"Direct key lookup: {key_count} keys")
    else:
        base_time_ms = 5.0 * key_count  # ~5ms per key for index scan
        factors.append(f"Index scan required: {key_count} keys")
    
    # Additional time for JOINs
    if 'JOIN' in query_upper:
        if 'ON KEYS' in query_upper:
            base_time_ms += 2.0 * key_count  # Key-based JOIN
            factors.append("Key-based JOIN operation")
        else:
            base_time_ms += 10.0 * key_count  # Index-based JOIN
            factors.append("Index-based JOIN (slower)")
    
    # Additional time for aggregations
    if any(agg in query_upper for agg in ['COUNT(', 'SUM(', 'AVG(', 'MIN(', 'MAX(']):
        base_time_ms += 5.0 * key_count
        factors.append("Aggregation functions detected")
    
    # Additional time for ORDER BY
    if 'ORDER BY' in query_upper:
        base_time_ms += key_count * 0.1  # Sorting overhead
        factors.append("Sorting overhead (ORDER BY)")
    
    # Additional time for DISTINCT
    if 'DISTINCT' in query_upper:
        base_time_ms += key_count * 0.2
        factors.append("DISTINCT processing")
    
    # DELETE operations are slightly faster
    if query_type == 'DELETE':
        base_time_ms *= 0.8
        factors.append("DELETE operation (faster than SELECT)")
    
    # Determine complexity
    complexity_score = 0
    if 'JOIN' in query_upper:
        complexity_score += 2
    if 'SUBQUERY' in query_upper or query_upper.count('SELECT') > 1:
        complexity_score += 2
    if any(agg in query_upper for agg in ['COUNT(', 'SUM(', 'AVG(', 'MIN(', 'MAX(']):
        complexity_score += 1
    if 'ORDER BY' in query_upper or 'GROUP BY' in query_upper:
        complexity_score += 1
    
    if complexity_score == 0:
        complexity = "Simple"
    elif complexity_score <= 2:
        complexity = "Moderate"
    else:
        complexity = "Complex"
    
    # Adjust estimate based on key count
    if key_count > 1000:
        base_time_ms *= 1.2  # Network overhead for large batches
        factors.append("Large batch size multiplier (1.2x)")
    
    return PerformanceEstimate(
        estimated_time_ms=round(base_time_ms, 2),
        complexity=complexity,
        key_count=key_count,
        factors=factors
    )


def format_entity_name(entity_name: str) -> str:
    """
    Ensure entity name is prefixed with a backtick.
    """
    if not entity_name:
        return entity_name
    return entity_name if entity_name.startswith("`") else f"`{entity_name}"


# ===============================
# MAIN EXECUTION
# ===============================


def main():
    """
    Main execution function.
    """
    config = Config()
    if not ensure_folders(config):
        print(color("\nERROR: Please ensure all folders exist before running.", ANSI_RED))
        sys.exit(1)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = config.log_file_location / f"execution_{timestamp}.log"
    log_handle = log_file.open("w", encoding="utf-8")
    original_stdout = sys.stdout
    original_stderr = sys.stderr
    log_stream = StripAnsiStream(log_handle)
    sys.stdout = Tee(original_stdout, log_stream)
    sys.stderr = Tee(original_stderr, log_stream)

    try:
        overall_start = time.time()
        banner("COUCHBASE N1QL QUERY GENERATOR")
        print(color(f"Log file: {log_file}", ANSI_GREEN))

        # Get Excel files from input folder
        step_title(f"Scanning input folder: {config.input_folder}")
        excel_files = get_excel_files(config.input_folder)

        if not excel_files:
            print(color("No Excel files found in input folder.", ANSI_YELLOW))
            sys.exit(0)

        print(color(f"Found {len(excel_files)} Excel file(s):", ANSI_CYAN))
        for i, file in enumerate(excel_files, 1):
            print(color(f"  {i}. {file.name}", ANSI_CYAN))

        report_rows: List[Dict[str, Any]] = []
        output_workbook = Workbook()
        sheet_names: set = set()

        # Process each Excel file
        total_files = len(excel_files)
        for idx, excel_file in enumerate(excel_files, 1):
            print(color(f"\n[File {idx}/{total_files}] {excel_file.name}", ANSI_MAGENTA))
            try:
                report_row, oc_output_df, rtb_output_df = process_excel_file(excel_file, config)

                base_name = excel_file.stem
                oc_sheet = make_sheet_title(base_name, "oc", sheet_names)
                rtb_sheet = make_sheet_title(base_name, "rtb", sheet_names)
                write_df_to_sheet(output_workbook, oc_sheet, oc_output_df)
                write_df_to_sheet(output_workbook, rtb_sheet, rtb_output_df)

                report_row["oc_sheet"] = oc_sheet
                report_row["rtb_sheet"] = rtb_sheet
                report_rows.append(report_row)
            except Exception as e:
                print(color(f"\nERROR processing {excel_file.name}: {str(e)}", ANSI_RED))
                import traceback
                traceback.print_exc()
                report_rows.append({
                    "file_name": excel_file.name,
                    "status": "Error",
                    "total_records": 0,
                    "oc_rows": 0,
                    "rtb_rows": 0,
                    "skipped_rows": 0,
                    "oc_groups": 0,
                    "rtb_groups": 0,
                    "temp_file": "",
                    "oc_sheet": "",
                    "rtb_sheet": "",
                    "read_time_s": 0.0,
                    "split_time_s": 0.0,
                    "query_gen_time_s": 0.0,
                    "file_time_s": 0.0,
                    "error": str(e),
                })
                continue

        if "Sheet" in output_workbook.sheetnames and len(output_workbook.sheetnames) > 1:
            default_sheet = output_workbook["Sheet"]
            if default_sheet.max_row == 1 and default_sheet.max_column == 1 and default_sheet["A1"].value is None:
                output_workbook.remove(default_sheet)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        combined_output_file = config.output_folder / f"n1ql_output_{timestamp}.xlsx"
        output_workbook.save(combined_output_file)

        report_file = config.execution_report_path / f"execution_report_{timestamp}.xlsx"
        save_execution_report(report_rows, report_file)

        print_summary_table(report_rows)
        banner("PROCESSING COMPLETE")
        print(color(f"Combined output file saved to: {combined_output_file}", ANSI_GREEN))
        print(color(f"Temporary files saved to: {config.temp_excel_path}", ANSI_GREEN))
        print(color(f"Execution report saved to: {config.execution_report_path}", ANSI_GREEN))
        print(color(f"Log file saved to: {log_file}", ANSI_GREEN))
        print(color(f"Total runtime: {time.time() - overall_start:.2f}s", ANSI_CYAN))
        print()
    finally:
        sys.stdout = original_stdout
        sys.stderr = original_stderr
        log_handle.close()


if __name__ == "__main__":
    main()

